import numpy as np
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, WebDriverException, NoSuchElementException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
import openpyxl as xl
from openpyxl import load_workbook

def stop(time_sleep):
    time.sleep(np.random.rand() + time_sleep)

def append_value(info, elems):
    for e in elems:
        info.append(e.text)

def add_name(info, elems):
    if len(elems) == 0:
        info.extend(['-','-','-'])

    elif len(elems) == 1:
        append_value(info, elems)
        info.extend(['-','-'])

    elif len(elems) == 2:
        append_value(info, elems)
        info.append(['-'])

    else: 
        elems = elems[:3]
        append_value(info, elems)
        info.append(['-'])

def get_values_table(table_xpath, row_i, col_i):
    table_id = driver.find_element(By.XPATH, table_xpath)
    rows = table_id.find_elements(By.TAG_NAME, "tr")
    total_values = []
    for i in row_i: 
        row_values = []
        for j in col_i:
            col_value = rows[i].find_elements(By.TAG_NAME, "td")[j].text
            row_values.append(col_value)
        total_values.append(row_values)
    return total_values

def click(driver, xpath):
    ActionChains(driver).move_to_element(WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, xpath)))).click().perform()

main_path = '/html/body/div[2]/div[2]/div/div[2]/div[1]/div[2]'
table_xpath = '//html/body/div[2]/div[2]/div/div[2]/div[2]/div[2]/div/table/tbody/tr'
regist_no = '/div[1]/div[1]/h3'                                        #'เลขทะเบียนนิติบุคคล'
juristic_name = '/div[1]/div[1]/h4'                                    #'ชื่อนิติบุคคล'
regist_type = '/div[2]/div[1]/div[1]/div/div/div/div[2]'               #'ประเภทนิติบุคคล'
status = '/div[2]/div[1]/div[1]/div/div/div/div[4]'                    #'สถานะนิติบุคคล'
regist_date = '/div[2]/div[1]/div[1]/div/div/div/div[6]'               #'วันที่จดทะเบียนจัดตั้ง'
regist_capital = '/div[2]/div[1]/div[1]/div/div/div/div[8]'            #'ทุนจดทะเบียน'
buss_type = '/div[2]/div[1]/div[1]/div/div/div/div[12]'                #'กลุ่มธุรกิจ'
buss_size = '/div[2]/div[1]/div[1]/div/div/div/div[14]'                #'ขนาดธุรกิจ'
buss_address = '/div[2]/div[1]/div[1]/div/div/div/div[18]'             #'ที่ตั้งสำนักงานแห่งใหญ่'
buss_type_regist = '/div[2]/div[1]/div[3]/div[2]/div[1]/div/div[2]'    #'ประเภทธุรกิจตอนจดทะเบียน'
buss_purpose_regist = '/div[2]/div[1]/div[3]/div[2]/div[1]/div/div[4]' #'วัตถุประสงค์ตอนจดทะเบียน
buss_type_last_state = '/div[2]/div[1]/div[3]/div[2]/div[2]/div/div[2]'#'ประเภทธุรกิจที่ส่งงบการเงินปีล่าสุด'
buss_purpose_state = '/div[2]/div[1]/div[3]/div[2]/div[2]/div/div[4]'  #'วัตถุประสงค์ที่ส่งงบการเงินปีล่าสุด'
all_xpath = [regist_no, juristic_name, regist_type, status, regist_date, regist_capital, buss_type, buss_size, buss_address, 
            buss_type_regist, buss_purpose_regist, buss_type_last_state, buss_purpose_state]
compare_year = '/html/body/div[2]/div[2]/div/div[2]/div[3]/div[2]/div[2]/div[1]/div/div/div[2]/div/div/label[1]/div/span[1]'
compare_bussiness = '/html/body/div[2]/div[2]/div/div[2]/div[3]/div[2]/div[2]/div[1]/div/div/div[2]/div/div/label[2]/div/span[1]'

excel_path1 = r'C:\Users\patcharapol.y\Desktop\Projects\New folder\VS code\dbd_web_scraping\DBD_scraping3.xlsx'
df_root = pd.read_excel(excel_path1, sheet_name='Sheet1')
tax_ids = df_root['เลขนิติบุลคล'].values

# Open Webdriver
driver = webdriver.Chrome()
driver.set_window_size(1024, 600)
driver.maximize_window()
driver.get('https://datawarehouse.dbd.go.th/index#_=_')
driver.find_element(By.XPATH, '//button[@class="btn btn-secondary"]').click()

for i, id in enumerate(tax_ids):
    try:
        print(f'Tax id: {i} {id}')
        wb = load_workbook(excel_path1)
        wb.create_sheet(str(id))
        ws = wb[str(id)]
        ws.cell(i+1, 1).value = id
        
        # Search ID
        elem = driver.find_element(By.XPATH, '//input[@name="textSearch"]')
        elem.clear()
        elem.send_keys(str(id))
        search_button = driver.find_element(By.XPATH,  '//span[@id="searchicon"]')
        ActionChains(driver).move_to_element(search_button).click().perform()
        stop(3)
        
        # Find click area elements
        click(driver, table_xpath + f'[1]/td[2]')
        stop(1)
        
        # Get Company Profile
        profile = []
        for xpath in all_xpath:
            elem = driver.find_element(By.XPATH, main_path + xpath).text
            profile.append(elem)
        
        # Get Committee names
        directors = driver.find_elements(By.XPATH, main_path + '/div[2]/div[1]/div[2]/div/div/ol')
        authorized_directors = driver.find_elements(By.XPATH, main_path + '/div[2]/div[1]/div[3]/div[1]/div')
        add_name(profile, directors)
        add_name(profile, authorized_directors)
        
        # Get statement years
        elems = driver.find_elements(By.XPATH, main_path + '/div[2]/div[1]/div[1]/div/div/div/div[16]/span')
        statement_years = [e.text for e in elems]
        profile.append(statement_years)
        for i_col in range(len(profile)):
            ws.cell(2, i_col + 1).value = str(profile[i_col])
            
        # click last year statement
        click(driver, '/html/body/div[2]/div[2]/div/div[2]/div[1]/div[2]/div[2]/div[1]/div[1]/div/div/div/div[16]/span[1]')
        stop(1)
        
        # Get Financial Status
        table_path = '/html/body/div[2]/div[2]/div/div[2]/div[3]/div[2]/div[2]/div[2]/div[2]/div/div[1]/div/table/tbody'
        financial_status = get_values_table(table_xpath=table_path, row_i=range(11), col_i=range(0, 9, 2))
        click(driver, compare_bussiness)
        stop(1)
        compare_financial_status = get_values_table(table_xpath=table_path, row_i=range(11), col_i=[1]) # bug ออกมาก 11 list in a list
        for i, row in enumerate(financial_status):
            row.append(compare_financial_status[i][0])
        for i_row in range(len(financial_status)):
            for i_col in range(len(financial_status[0])):
                ws.cell(i_row + 1 + 5, i_col + 1).value = str(financial_status[i_row][i_col])
                
        # Get Profit Loss 
        profit_loss_xpath = '/html/body/div[2]/div[2]/div/div[2]/div[3]/div[2]/div[2]/div[1]/div/div/div[1]/div/div[2]/span[2]'
        click(driver, profit_loss_xpath)
        stop(1)
        click(driver, compare_year)
        stop(1)
        profit_loss = get_values_table(table_xpath=table_path, row_i=range(10), col_i=range(0, 9, 2))
        click(driver, compare_bussiness)
        stop(1)
        compare_profit_loss = get_values_table(table_xpath=table_path, row_i=range(10), col_i=[1])
        for i, row in enumerate(profit_loss):
            row.append(compare_profit_loss[i][0])
        for i_row in range(len(profit_loss)):
            for i_col in range(len(profit_loss[0])):
                ws.cell(i_row + 1 + 17, i_col + 1).value = str(profit_loss[i_row][i_col])
                
        # get last year statement date
        driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.CONTROL + Keys.HOME)
        stop(1)
        span = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[1]/ul/li[2]/span/span')
        ActionChains(driver).move_to_element(span).perform()
        history_statement_xpath = '/html/body/div[2]/div[2]/div/div[1]/ul/li[2]/ul/li[3]/a'
        click(driver, history_statement_xpath)
        stop(1)
        last_statement = get_values_table(table_xpath='/html/body/div[2]/div[2]/div/div[2]/div[4]/div[2]/div[2]/div[1]/div/div/div/div/table/tbody', row_i=[-1], col_i=[0, 2])
        for i_row in range(len(last_statement)):
            for i_col in range(len(last_statement[0])):
                ws.cell(i_row + 1 + 3, i_col + 1).value = str(last_statement[i_row][i_col])
                
        # get stakeholders
        driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.CONTROL + Keys.HOME)
        stop(1)
        investors_xpath = '/html/body/div[2]/div[2]/div/div[1]/ul/li[3]/a'
        click(driver, investors_xpath)
        stop(1)
        table_path='/html/body/div[2]/div[2]/div/div[2]/div[5]/div[2]/div[2]/div[1]/div/div[2]/div/table/tbody'
        investors = get_values_table(table_xpath=table_path, row_i=range(2), col_i=range(1,4))
        for i_row in range(len(investors)):
            for i_col in range(len(investors[0])):
                ws.cell(i_row + 1 + 29, i_col + 1).value = str(investors[i_row][i_col])
    except:
        pass
    wb.save(excel_path1)


